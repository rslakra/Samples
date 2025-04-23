import json
import os

import openpyxl


# Press the green button in the gutter to run the script.
# starting point

# writes the json list into given file.
def write_json(output_file_path, delimiter, json_list):
    # print(f'write_json({output_file_path}, {json_list})')
    print(f'write_json({output_file_path})')
    # open file to write into
    with open(jsonFilePath, "w") as file:
        file.write(json_list)

    # close json file
    file.close()


if __name__ == '__main__':
    print()
    userHome = os.getenv("HOME")
    print("userHome: ", userHome)
    downloadFolderName = "/Downloads/"
    fileName = "Prepay Updates July 2021"
    filePath = userHome + downloadFolderName + fileName + ".xlsx"
    print("filePath: ", filePath)
    # csvFile = open(csvFilePath, "r")
    print()
    # print(csvFile.readline())

    workBook = openpyxl.load_workbook(filePath)
    workBook.active = workBook['AMEND']
    sheet = workBook.active
    print(sheet)
    maxRows = sheet.max_row
    maxCols = sheet.max_column
    print("maxRows:", maxRows)
    print("maxCols:", maxCols)
    # # Loop will print all columns name
    # for i in range(1, maxCols + 1):
    #     cell_obj = sheet.cell(row=1, column=i)
    #     print(cell_obj.value)

    # # print all row/column values
    # for row in range(1, maxRows + 1):
    #     for col in range(1, maxCols + 1):
    #         cell = sheet.cell(row=row, column=col)
    #         print(cell.value);

    # Loop will print all values of 2nd column
    crmCustomerIds = []
    jsonResult = []
    for row in range(2, maxRows + 1):
        # cell = sheet.cell(row=row, column=2)
        # values.append(cell.value)
        # print(cell.value)
        crmCustomerId = sheet.cell(row=row, column=2).value
        vatId = sheet.cell(row=row, column=4).value
        advertiserId = sheet.cell(row=row, column=9).value
        # print("crmCustomerId:", crmCustomerId, ", advertiserId:", advertiserId, ", vatId:", vatId)
        # values.append("'" + cell.value + "', ")
        crmCustomerIds.append(crmCustomerId)
        vatDict = {}
        vatDict["id"] = advertiserId
        vatDict["crmCustomerId"] = crmCustomerId
        vatDict["vatId"] = vatId
        vatDict["vatOptOut"] = "false"
        # print(vatDict)
        # Add record to dictionary
        jsonResult.append(vatDict)

    # print(jsonResult)
    # print("crmCustomerIds:", crmCustomerIds)
    print("jsonResult:", len(jsonResult))

# open json file
jsonFilePath = userHome + downloadFolderName + fileName + "Update.json"
print("jsonFilePath: ", jsonFilePath)
jsonObject = json.dumps(jsonResult, indent=4)
# print(jsonObject)
write_json(jsonFilePath, "\n", jsonObject)
print()

# # open json file
# jsonFilePath = userHome + downloadFolderName + fileName + "-ids.json"
# print("jsonFilePath: ", jsonFilePath)
# write_json(jsonFilePath, "", ids)

print()
